"""可追溯的绿的谐波简化财务模型。

模型定位是 P0 原型：用本地 CSV 输入计算三种经营情景，不代表正式投资预测。
每条输入都带有 ``input_type``：

* ``historical``：历史披露锚点，只作展示或校验；
* ``assumption``：人工设定的模型假设；
* ``calculated``：运行时产生的结果，不能写回输入表。

计算链：销量 × ASP → 收入 → COGS → 毛利 → EBITDA → EBIT → FCF → DCF。
金额统一为十亿元人民币（bn CNY），销量统一为万台，ASP/成本统一为元/台。
"""

from __future__ import annotations

import argparse
import csv
import json
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Sequence


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT_PATH = REPO_ROOT / "data" / "processed" / "green_harmonic_model_inputs.csv"
DEFAULT_MODEL_PATH = REPO_ROOT / "data" / "processed" / "green_harmonic_model.xlsx"
DEFAULT_JSON_PATH = REPO_ROOT / "data" / "processed" / "green_harmonic_model_results.json"

FORECAST_YEARS = (2025, 2026, 2027)
REQUIRED_METRICS = (
    "harmonic_units_wan",
    "joint_units_wan",
    "harmonic_asp_yuan",
    "joint_asp_yuan",
    "harmonic_unit_cost_yuan",
    "joint_unit_cost_yuan",
    "selling_expense_rate",
    "admin_expense_rate",
    "rd_expense_rate",
    "depreciation_rate",
    "tax_rate",
    "capex_rate",
    "nwc_rate",
    "discount_rate",
    "terminal_growth_rate",
)


@dataclass(frozen=True)
class ModelInput:
    metric: str
    year: str
    value: float
    unit: str
    input_type: str
    source: str
    source_locator: str
    notes: str = ""


class FinancialModelInputs:
    """带来源元数据的模型输入集合。"""

    def __init__(self, rows: Iterable[ModelInput]):
        self.rows = list(rows)
        self._values = {(row.metric, row.year): row for row in self.rows}

    @classmethod
    def from_csv(cls, path: Path | str) -> "FinancialModelInputs":
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"模型输入不存在：{path}")
        required_columns = {
            "metric", "year", "value", "unit", "input_type",
            "source", "source_locator", "notes",
        }
        rows: List[ModelInput] = []
        seen_keys = set()
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            missing = required_columns - set(reader.fieldnames or [])
            if missing:
                raise ValueError(f"模型输入缺少字段：{sorted(missing)}")
            for line_no, raw in enumerate(reader, start=2):
                metric = raw["metric"].strip()
                year = raw["year"].strip()
                if not metric or not year:
                    raise ValueError(f"第 {line_no} 行 metric/year 不能为空")
                if (metric, year) in seen_keys:
                    raise ValueError(f"第 {line_no} 行 metric/year 重复：{metric}/{year}")
                source = raw["source"].strip()
                source_locator = raw["source_locator"].strip()
                if not source or not source_locator:
                    raise ValueError(f"第 {line_no} 行必须填写 source 和 source_locator")
                try:
                    value = float(raw["value"])
                except (TypeError, ValueError) as exc:
                    raise ValueError(f"第 {line_no} 行 value 不是数字") from exc
                if not math.isfinite(value):
                    raise ValueError(f"第 {line_no} 行 value 不是有限数字")
                input_type = raw["input_type"].strip().lower()
                if input_type not in {"historical", "assumption"}:
                    raise ValueError(
                        f"第 {line_no} 行 input_type 必须是 historical 或 assumption"
                    )
                rows.append(ModelInput(
                    metric=metric,
                    year=year,
                    value=value,
                    unit=raw["unit"].strip(),
                    input_type=input_type,
                    source=source,
                    source_locator=source_locator,
                    notes=raw.get("notes", "").strip(),
                ))
                seen_keys.add((metric, year))
        return cls(rows)

    def get(self, metric: str, year: int | str) -> ModelInput:
        key = (metric, str(year))
        try:
            return self._values[key]
        except KeyError as exc:
            raise KeyError(f"缺少模型输入：{metric}/{year}") from exc

    def value(self, metric: str, year: int | str) -> float:
        return self.get(metric, year).value

    def validate(self, years: Sequence[int] = FORECAST_YEARS) -> None:
        missing = []
        for metric in REQUIRED_METRICS:
            year = "valuation" if metric in {"discount_rate", "terminal_growth_rate"} else None
            if year is not None:
                if (metric, year) not in self._values:
                    missing.append(f"{metric}/{year}")
            else:
                for forecast_year in years:
                    if (metric, str(forecast_year)) not in self._values:
                        missing.append(f"{metric}/{forecast_year}")
        if missing:
            raise ValueError(f"模型输入不完整：{', '.join(missing)}")

    def as_records(self) -> List[Dict[str, Any]]:
        return [
            {
                "metric": row.metric,
                "year": row.year,
                "value": row.value,
                "unit": row.unit,
                "input_type": row.input_type,
                "source": row.source,
                "source_locator": row.source_locator,
                "notes": row.notes,
            }
            for row in self.rows
        ]


SCENARIOS: Mapping[str, Mapping[str, float]] = {
    "base": {
        "volume_multiplier": 1.00,
        "asp_multiplier": 1.00,
        "unit_cost_multiplier": 1.00,
        "opex_multiplier": 1.00,
    },
    "upside": {
        "volume_multiplier": 1.10,
        "asp_multiplier": 1.03,
        "unit_cost_multiplier": 0.97,
        "opex_multiplier": 0.98,
    },
    "downside": {
        "volume_multiplier": 0.90,
        "asp_multiplier": 0.97,
        "unit_cost_multiplier": 1.05,
        "opex_multiplier": 1.03,
    },
}


def _money_from_units(units_wan: float, price_yuan: float) -> float:
    """万台 × 元/台转换为十亿元人民币。"""
    return units_wan * price_yuan / 100_000.0


def _round_values(value: Any) -> Any:
    if isinstance(value, float):
        return round(value, 8)
    if isinstance(value, dict):
        return {key: _round_values(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_round_values(item) for item in value]
    return value


def run_model(
    inputs: FinancialModelInputs,
    scenario: str = "base",
    years: Sequence[int] = FORECAST_YEARS,
) -> Dict[str, Any]:
    """运行一个情景并返回可序列化结果。"""
    if scenario not in SCENARIOS:
        raise ValueError(f"未知情景：{scenario}，可选：{sorted(SCENARIOS)}")
    inputs.validate(years)
    settings = SCENARIOS[scenario]
    projections: List[Dict[str, Any]] = []
    previous_nwc = 0.0

    for year in years:
        volume_factor = settings["volume_multiplier"]
        asp_factor = settings["asp_multiplier"]
        cost_factor = settings["unit_cost_multiplier"]
        opex_factor = settings["opex_multiplier"]

        harmonic_units = inputs.value("harmonic_units_wan", year) * volume_factor
        joint_units = inputs.value("joint_units_wan", year) * volume_factor
        harmonic_asp = inputs.value("harmonic_asp_yuan", year) * asp_factor
        joint_asp = inputs.value("joint_asp_yuan", year) * asp_factor
        harmonic_cost = inputs.value("harmonic_unit_cost_yuan", year) * cost_factor
        joint_cost = inputs.value("joint_unit_cost_yuan", year) * cost_factor

        harmonic_revenue = _money_from_units(harmonic_units, harmonic_asp)
        joint_revenue = _money_from_units(joint_units, joint_asp)
        harmonic_cogs = _money_from_units(harmonic_units, harmonic_cost)
        joint_cogs = _money_from_units(joint_units, joint_cost)
        revenue = harmonic_revenue + joint_revenue
        cogs = harmonic_cogs + joint_cogs
        gross_profit = revenue - cogs

        selling = inputs.value("selling_expense_rate", year) * opex_factor
        admin = inputs.value("admin_expense_rate", year) * opex_factor
        rd = inputs.value("rd_expense_rate", year) * opex_factor
        opex = revenue * (selling + admin + rd)
        ebitda = gross_profit - opex
        depreciation = revenue * inputs.value("depreciation_rate", year)
        ebit = ebitda - depreciation
        tax = max(ebit, 0.0) * inputs.value("tax_rate", year)
        nopat = ebit - tax
        capex = revenue * inputs.value("capex_rate", year)
        nwc = revenue * inputs.value("nwc_rate", year)
        change_nwc = nwc - previous_nwc
        fcf = nopat + depreciation - capex - change_nwc
        previous_nwc = nwc

        projections.append({
            "year": year,
            "harmonic_units_wan": harmonic_units,
            "joint_units_wan": joint_units,
            "harmonic_revenue_bn": harmonic_revenue,
            "joint_revenue_bn": joint_revenue,
            "revenue_bn": revenue,
            "harmonic_cogs_bn": harmonic_cogs,
            "joint_cogs_bn": joint_cogs,
            "cogs_bn": cogs,
            "gross_profit_bn": gross_profit,
            "gross_margin": gross_profit / revenue if revenue else 0.0,
            "opex_bn": opex,
            "ebitda_bn": ebitda,
            "ebitda_margin": ebitda / revenue if revenue else 0.0,
            "depreciation_bn": depreciation,
            "ebit_bn": ebit,
            "tax_bn": tax,
            "capex_bn": capex,
            "change_nwc_bn": change_nwc,
            "fcf_bn": fcf,
        })

    discount_rate = inputs.value("discount_rate", "valuation")
    terminal_growth = inputs.value("terminal_growth_rate", "valuation")
    if discount_rate <= terminal_growth:
        raise ValueError("discount_rate 必须大于 terminal_growth_rate")
    discounted_fcfs = []
    for index, projection in enumerate(projections, start=1):
        discounted_fcfs.append(projection["fcf_bn"] / ((1 + discount_rate) ** index))
    terminal_fcf = projections[-1]["fcf_bn"] * (1 + terminal_growth)
    terminal_value = terminal_fcf / (discount_rate - terminal_growth)
    discounted_terminal = terminal_value / ((1 + discount_rate) ** len(projections))
    enterprise_value = sum(discounted_fcfs) + discounted_terminal

    result = {
        "model_name": "green_harmonic_simplified_financial_model",
        "model_status": "prototype_scenario_not_investment_recommendation",
        "scenario": scenario,
        "currency": "bn_cny",
        "years": list(years),
        "scenario_settings": dict(settings),
        "projections": projections,
        "valuation": {
            "discount_rate": discount_rate,
            "terminal_growth_rate": terminal_growth,
            "discounted_fcfs_bn": discounted_fcfs,
            "terminal_value_bn": terminal_value,
            "discounted_terminal_value_bn": discounted_terminal,
            "enterprise_value_bn": enterprise_value,
        },
        "traceability": {
            "input_types": {
                "historical": sum(row.input_type == "historical" for row in inputs.rows),
                "assumption": sum(row.input_type == "assumption" for row in inputs.rows),
                "calculated": len(projections),
            },
            "historical_inputs_used_for_forecast": False,
            "note": "历史披露锚点与人工假设保留在 Inputs 表；预测结果由本模型计算。",
        },
    }
    return _round_values(result)


def run_all_scenarios(inputs: FinancialModelInputs) -> Dict[str, Dict[str, Any]]:
    return {name: run_model(inputs, name) for name in SCENARIOS}


def export_json(results: Mapping[str, Dict[str, Any]], path: Path | str) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def export_excel(
    inputs: FinancialModelInputs,
    results: Mapping[str, Dict[str, Any]],
    path: Path | str,
) -> Path:
    """输出包含输入溯源、三情景和汇总的 xlsx。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:  # pragma: no cover - 依赖缺失时的清晰提示
        raise RuntimeError("输出 Excel 需要 openpyxl，请先安装 requirements.txt") from exc

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    input_sheet = workbook.active
    input_sheet.title = "Inputs"
    input_headers = [
        "metric", "year", "value", "unit", "input_type", "source",
        "source_locator", "notes",
    ]
    input_sheet.append(input_headers)
    for cell in input_sheet[1]:
        cell.font = Font(bold=True)
    for row in inputs.as_records():
        input_sheet.append([row[header] for header in input_headers])
    input_sheet.freeze_panes = "A2"

    summary = workbook.create_sheet("Summary")
    summary.append(["scenario", "enterprise_value_bn", "2027_revenue_bn", "2027_fcf_bn", "model_status"])
    for cell in summary[1]:
        cell.font = Font(bold=True)
    for name, result in results.items():
        last = result["projections"][-1]
        summary.append([
            name,
            result["valuation"]["enterprise_value_bn"],
            last["revenue_bn"],
            last["fcf_bn"],
            result["model_status"],
        ])

    metric_order = [
        "harmonic_units_wan", "joint_units_wan", "revenue_bn", "cogs_bn",
        "gross_profit_bn", "gross_margin", "opex_bn", "ebitda_bn",
        "ebitda_margin", "ebit_bn", "tax_bn", "capex_bn", "change_nwc_bn", "fcf_bn",
    ]
    for name, result in results.items():
        sheet = workbook.create_sheet(name.title())
        sheet.append(["metric", *[str(year) for year in result["years"]], "logic"])
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for metric in metric_order:
            values = [projection[metric] for projection in result["projections"]]
            logic = "calculated; see src/financial_model.py"
            sheet.append([metric, *values, logic])
        sheet.append([])
        sheet.append(["enterprise_value_bn", result["valuation"]["enterprise_value_bn"], "", "", "DCF"])
        sheet.freeze_panes = "A2"

    workbook.save(path)
    return path


def main() -> None:
    parser = argparse.ArgumentParser(description="绿的谐波可追溯简化财务模型")
    parser.add_argument("--inputs", default=str(DEFAULT_INPUT_PATH))
    parser.add_argument("--xlsx", default=str(DEFAULT_MODEL_PATH))
    parser.add_argument("--json", default=str(DEFAULT_JSON_PATH))
    args = parser.parse_args()

    inputs = FinancialModelInputs.from_csv(args.inputs)
    results = run_all_scenarios(inputs)
    export_json(results, args.json)
    export_excel(inputs, results, args.xlsx)
    base = results["base"]
    print(f"模型已完成：base enterprise value={base['valuation']['enterprise_value_bn']:.4f} bn CNY")
    print(f"JSON：{args.json}")
    print(f"Excel：{args.xlsx}")


if __name__ == "__main__":
    main()
